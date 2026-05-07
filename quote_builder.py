import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

# ── COLOURS ───────────────────────────────────────────────────────────────────
BLACK   = "1A1A1A"
ORANGE  = "C85A00"
WHITE   = "FFFFFF"
LGREY   = "F5F5F5"
LORANGE = "FFF3E0"
MGREY   = "CCCCCC"

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def font(bold=False, color=BLACK, size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)

def border():
    s = Side(style="thin", color=MGREY)
    return Border(left=s, right=s, top=s, bottom=s)

def search_books(query):
    conn = sqlite3.connect("books.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    # Search by ISBN or title
    cursor.execute("""
        SELECT isbn, title, grade, language, price, publisher 
        FROM books 
        WHERE isbn LIKE ? OR title LIKE ?
        LIMIT 20
    """, (f"%{query}%", f"%{query}%"))
    results = cursor.fetchall()
    conn.close()
    return [dict(r) for r in results]

    BLACK = "1A1A1A"
ORANGE = "C85A00"
WHITE = "FFFFFF"
LGREY = "F5F5F5"
MGREY = "CCCCCC"

def fill(h): return PatternFill("solid", start_color=h, fgColor=h)
def bfont(bold=False, color=BLACK, size=10): return Font(name="Arial", bold=bold, color=color, size=size)
def bdr():
    s = Side(style="thin", color=MGREY)
    return Border(left=s, right=s, top=s, bottom=s)

def build_quote(items, school, ref_no):
    wb = Workbook()
    ws = wb.active
    ws.title = "QUOTATION"

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 16

    ws.merge_cells("A1:F1")
    ws["A1"] = "KALIKENG TRADING AND PROJECTS CC"
    ws["A1"].font = bfont(bold=True, size=16, color=WHITE)
    ws["A1"].fill = fill(BLACK)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    ws["A2"] = "Reg No: 2010/007041/23  |  VAT No: 4340257320  |  Tel: 073 223 9762  |  kalikengtrading@gmail.com"
    ws["A2"].font = bfont(size=9, color=ORANGE)
    ws["A2"].fill = fill(BLACK)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    ws.merge_cells("A3:F3")
    ws["A3"] = "B-BBEE LEVEL 1  |  100% BLACK OWNED  |  135% PROCUREMENT RECOGNITION"
    ws["A3"].font = bfont(bold=True, size=9, color=WHITE)
    ws["A3"].fill = fill(BLACK)
    ws["A3"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[3].height = 16

    for col in range(1, 7):
        ws.cell(row=4, column=col).fill = fill(ORANGE)
    ws.row_dimensions[4].height = 4
    ws.row_dimensions[5].height = 6

    ws.row_dimensions[6].height = 18
    ws.row_dimensions[7].height = 18

    ws.merge_cells("A6:B6")
    ws["A6"] = "QUOTATION NO:"
    ws["A6"].font = bfont(bold=True, color=ORANGE, size=9)
    ws["A6"].fill = fill(LGREY)
    ws["A6"].border = bdr()

    ws.merge_cells("C6:D6")
    ws["C6"] = ref_no
    ws["C6"].font = bfont(size=9)
    ws["C6"].fill = fill(WHITE)
    ws["C6"].border = bdr()

    ws["E6"] = "DATE:"
    ws["E6"].font = bfont(bold=True, color=ORANGE, size=9)
    ws["E6"].fill = fill(LGREY)
    ws["E6"].border = bdr()

    ws["F6"] = str(date.today().strftime("%d %B %Y"))
    ws["F6"].font = bfont(size=9)
    ws["F6"].fill = fill(WHITE)
    ws["F6"].border = bdr()

    ws.merge_cells("A7:B7")
    ws["A7"] = "SCHOOL / CLIENT:"
    ws["A7"].font = bfont(bold=True, color=ORANGE, size=9)
    ws["A7"].fill = fill(LGREY)
    ws["A7"].border = bdr()

    ws.merge_cells("C7:D7")
    ws["C7"] = school
    ws["C7"].font = bfont(size=9)
    ws["C7"].fill = fill(WHITE)
    ws["C7"].border = bdr()

    ws["E7"] = "PREPARED BY:"
    ws["E7"].font = bfont(bold=True, color=ORANGE, size=9)
    ws["E7"].fill = fill(LGREY)
    ws["E7"].border = bdr()

    ws["F7"] = "Sechaba Mofokeng"
    ws["F7"].font = bfont(size=9)
    ws["F7"].fill = fill(WHITE)
    ws["F7"].border = bdr()

    ws.row_dimensions[8].height = 6

    for col, hdr in zip(["A","B","C","D","E","F"],
                        ["#","ISBN","TITLE","UNIT PRICE","QTY","LINE TOTAL"]):
        c = ws[f"{col}9"]
        c.value = hdr
        c.font = bfont(bold=True, color=WHITE, size=9)
        c.fill = fill(BLACK)
        c.alignment = Alignment(horizontal="center")
        c.border = bdr()
    ws.row_dimensions[9].height = 18

    grand_total = 0
    for idx, item in enumerate(items):
        r = 10 + idx
        shade = LGREY if idx % 2 == 0 else WHITE
        line_total = float(item["price"]) * int(item["qty"])
        grand_total += line_total
        ws.row_dimensions[r].height = 18

        isbn_clean = str(item["isbn"]).replace(".0","")
        data = [("A", idx+1), ("B", isbn_clean), ("C", item["title"]),
                ("D", float(item["price"])), ("E", int(item["qty"])), ("F", line_total)]

        for col, val in data:
            c = ws[f"{col}{r}"]
            c.value = val
            c.font = bfont(size=9)
            c.fill = fill(shade)
            c.border = bdr()
            if col in ["D","F"]:
                c.number_format = 'R#,##0.00'
            if col in ["A","D","E","F"]:
                c.alignment = Alignment(horizontal="center")

    tr = 10 + len(items)
    ws.merge_cells(f"A{tr}:E{tr}")
    tc = ws[f"A{tr}"]
    tc.value = "GRAND TOTAL"
    tc.font = bfont(bold=True, color=WHITE, size=11)
    tc.fill = fill(BLACK)
    tc.alignment = Alignment(horizontal="right", vertical="center")
    tc.border = bdr()
    ws.row_dimensions[tr].height = 22

    gt = ws[f"F{tr}"]
    gt.value = grand_total
    gt.font = bfont(bold=True, color=WHITE, size=11)
    gt.fill = fill(ORANGE)
    gt.border = bdr()
    gt.number_format = 'R#,##0.00'

    fr = tr + 2
    ws.merge_cells(f"A{fr}:F{fr}")
    ft = ws[f"A{fr}"]
    ft.value = "Kalikeng Trading and Projects CC  |  Excellence in Service Delivery Since 2010  |  kalikengtrading@gmail.com"
    ft.font = bfont(size=8, color="555555")
    ft.alignment = Alignment(horizontal="center")

    filename = f"Quote_{ref_no}_{school.replace(' ','_')}.xlsx"
    wb.save(filename)
    return filename, grand_total


def main():
    print("\n" + "="*60)
    print("   KALIKENG TEXTBOOK & STATIONERY QUOTE BUILDER")
    print("="*60)

    school = input("\nEnter school/client name: ").strip()
    ref_no = input("Enter quotation number (e.g. KT-2026-001): ").strip()

    items = []
    print("\nSearch by ISBN or title. Type 'done' when finished.\n")

    while True:
        query = input("Search ISBN or title (or 'done'): ").strip()
        if query.lower() == "done":
            break
        if not query:
            continue

        results = search_books(query)
        if not results:
            print("  No results found. Try again.\n")
            continue

        print(f"\n  Found {len(results)} result(s):")
        for i, book in enumerate(results, 1):
            isbn_clean = str(book["isbn"]).replace(".0","")
            print(f"  [{i}] {isbn_clean} | {book['title'][:50]} | Gr{book['grade']} | R{book['price']} | {book['publisher']}")

        choice = input("\n  Select number (or Enter to skip): ").strip()
        if not choice:
            continue

        try:
            selected = results[int(choice) - 1]
            qty = input(f"  Quantity for '{selected['title'][:40]}': ").strip()
            qty = int(qty)
            if qty > 0:
                selected["qty"] = qty
                selected["isbn"] = str(selected["isbn"]).replace(".0","")
                items.append(selected)
                line = qty * selected["price"]
                print(f"  ✓ Added — Line total: R{line:,.2f}\n")
        except (ValueError, IndexError):
            print("  Invalid selection. Try again.\n")
            continue

    if not items:
        print("\nNo items added. Exiting.")
        return

    print(f"\nBuilding quote with {len(items)} items...")
    filename, total = build_quote(items, school, ref_no)
    print(f"\n{'='*60}")
    print(f"  QUOTE SAVED: {filename}")
    print(f"  GRAND TOTAL: R{total:,.2f}")
    print(f"{'='*60}\n")

if __name__ == "__main__":
    main()