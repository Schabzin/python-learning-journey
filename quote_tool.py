import tkinter as tk
from tkinter import ttk, messagebox
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

# ── COLOURS ───────────────────────────────────────────────────────────────────
BLACK = "1A1A1A"
ORANGE = "C85A00"
WHITE = "FFFFFF"
LGREY = "F5F5F5"
MGREY = "CCCCCC"

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def bfont(bold=False, color=BLACK, size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)

def bdr():
    s = Side(style="thin", color=MGREY)
    return Border(left=s, right=s, top=s, bottom=s)

def search_books(query):
    conn = sqlite3.connect("books.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute("""
    SELECT isbn, title, grade, subject, language, price, publisher 
    FROM books 
    WHERE REPLACE(isbn,'.0','') LIKE ? OR LOWER(title) LIKE LOWER(?)
    ORDER BY title LIMIT 30
""", (f"%{query}%", f"%{query.lower()}%"))
    results = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return results

def export_quote(items, school, ref_no):
    wb = Workbook()
    ws = wb.active
    ws.title = "QUOTATION"

    for col, w in [("A",20),("B",50),("C",8),("D",14),("E",14),("F",14),("G",12),("H",16)]:
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:H1")
    ws["A1"] = "KALIKENG TRADING AND PROJECTS CC"
    ws["A1"].font = bfont(bold=True, size=16, color=WHITE)
    ws["A1"].fill = fill(BLACK)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "Reg No: 2010/007041/23  |  VAT No: 4340257320  |  Tel: 073 223 9762  |  kalikengtrading@gmail.com"
    ws["A2"].font = bfont(size=9, color=ORANGE)
    ws["A2"].fill = fill(BLACK)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:H3")
    ws["A3"] = "B-BBEE LEVEL 1  |  100% BLACK OWNED  |  135% PROCUREMENT RECOGNITION"
    ws["A3"].font = bfont(bold=True, size=9, color=WHITE)
    ws["A3"].fill = fill(BLACK)
    ws["A3"].alignment = Alignment(horizontal="center")

    for col in range(1, 9):
        ws.cell(row=4, column=col).fill = fill(ORANGE)
    ws.row_dimensions[4].height = 4
    ws.row_dimensions[5].height = 6

    details = [
        (6, "A", "QUOTATION NO:", "B", ref_no, "E", "DATE:", "F", str(date.today().strftime("%d %B %Y"))),
        (7, "A", "SCHOOL / CLIENT:", "B", school, "E", "PREPARED BY:", "F", "Sechaba Mofokeng"),
    ]
    for row, l1, v1, c1, val1, l2, v2, c2, val2 in details:
        ws.row_dimensions[row].height = 18
        for col, val, bold in [(l1,v1,True),(c1,val1,False),(l2,v2,True),(c2,val2,False)]:
            c = ws[f"{col}{row}"]
            c.value = val
            c.font = bfont(bold=bold, color=ORANGE if bold else BLACK, size=9)
            c.fill = fill(LGREY if bold else WHITE)
            c.border = bdr()

    ws.row_dimensions[8].height = 6

    for col, hdr in zip(["A","B","C","D","E","F","G","H"],
                        ["ISBN","TITLE","GR","SUBJECT","LANGUAGE","PUBLISHER","UNIT PRICE","QTY"]):
        c = ws[f"{col}9"]
        c.value = hdr
        c.font = bfont(bold=True, color=WHITE, size=9)
        c.fill = fill(BLACK)
        c.alignment = Alignment(horizontal="center")
        c.border = bdr()

    ws.column_dimensions["I"].width = 16
    c = ws["I9"]
    c.value = "LINE TOTAL"
    c.font = bfont(bold=True, color=WHITE, size=9)
    c.fill = fill(BLACK)
    c.alignment = Alignment(horizontal="center")
    c.border = bdr()
    ws.row_dimensions[9].height = 18

    grand_total = 0
    for idx, item in enumerate(items):
        r = 10 + idx
        shade = LGREY if idx % 2 == 0 else WHITE
        line_total = float(item["price"]) * int(item["qty"])
        grand_total += line_total
        ws.row_dimensions[r].height = 16

        isbn_clean = str(item["isbn"]).replace(".0","")
        data = [("A",isbn_clean),("B",item["title"]),("C",str(item["grade"])),
                ("D",str(item["subject"])),("E",str(item["language"])),
                ("F",str(item["publisher"])),("G",float(item["price"])),
                ("H",int(item["qty"])),("I",line_total)]

        for col, val in data:
            c = ws[f"{col}{r}"]
            c.value = val
            c.font = bfont(size=9)
            c.fill = fill(shade)
            c.border = bdr()
            if col in ["G","I"]:
                c.number_format = 'R#,##0.00'
            if col in ["G","H","I"]:
                c.alignment = Alignment(horizontal="center")

    tr = 10 + len(items)
    ws.merge_cells(f"A{tr}:H{tr}")
    tc = ws[f"A{tr}"]
    tc.value = "GRAND TOTAL"
    tc.font = bfont(bold=True, color=WHITE, size=11)
    tc.fill = fill(BLACK)
    tc.alignment = Alignment(horizontal="right", vertical="center")
    tc.border = bdr()
    ws.row_dimensions[tr].height = 22

    gt = ws[f"I{tr}"]
    gt.value = grand_total
    gt.font = bfont(bold=True, color=WHITE, size=11)
    gt.fill = fill(ORANGE)
    gt.border = bdr()
    gt.number_format = 'R#,##0.00'

    fr = tr + 2
    ws.merge_cells(f"A{fr}:I{fr}")
    ft = ws[f"A{fr}"]
    ft.value = "Kalikeng Trading and Projects CC  |  Excellence in Service Delivery Since 2010  |  kalikengtrading@gmail.com"
    ft.font = bfont(size=8, color="555555")
    ft.alignment = Alignment(horizontal="center")

    filename = f"Quote_{ref_no}_{school.replace(' ','_')}.xlsx"
    wb.save(filename)
    return filename, grand_total


class QuoteTool:
    def __init__(self, root):
        self.root = root
        self.root.title("Kalikeng Quote Builder")
        self.root.geometry("1100x700")
        self.root.configure(bg="#1A1A1A")
        self.quote_items = []
        self.build_ui()

    def build_ui(self):
        # ── HEADER ────────────────────────────────────────────────────────────
        hdr = tk.Frame(self.root, bg="#1A1A1A", pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr, text="KALIKENG TRADING AND PROJECTS CC",
                 font=("Arial", 16, "bold"), fg="#C85A00", bg="#1A1A1A").pack()
        tk.Label(hdr, text="Textbook & Stationery Quote Builder",
                 font=("Arial", 10), fg="white", bg="#1A1A1A").pack()

        # Orange divider
        tk.Frame(self.root, bg="#C85A00", height=3).pack(fill="x")

        # ── MAIN AREA ─────────────────────────────────────────────────────────
        main = tk.Frame(self.root, bg="#f5f5f5")
        main.pack(fill="both", expand=True, padx=10, pady=10)

        # LEFT — Search panel
        left = tk.Frame(main, bg="white", relief="solid", bd=1)
        left.pack(side="left", fill="both", expand=True, padx=(0,5))

        tk.Label(left, text="SEARCH BOOKS", font=("Arial", 11, "bold"),
                 fg="#1A1A1A", bg="white").pack(pady=(10,5))

        search_frame = tk.Frame(left, bg="white")
        search_frame.pack(fill="x", padx=10)
        self.search_var = tk.StringVar()
        search_entry = tk.Entry(search_frame, textvariable=self.search_var,
                               font=("Arial", 11), width=30)
        search_entry.pack(side="left", fill="x", expand=True, ipady=5)
        search_entry.bind("<Return>", lambda e: self.do_search())
        tk.Button(search_frame, text="Search", command=self.do_search,
                 bg="#C85A00", fg="white", font=("Arial", 10, "bold"),
                 relief="flat", padx=10).pack(side="left", padx=(5,0))

        tk.Label(left, text="Search by ISBN or title — press Enter or click Search",
                 font=("Arial", 8), fg="#888", bg="white").pack()

        # Results list
        results_frame = tk.Frame(left, bg="white")
        results_frame.pack(fill="both", expand=True, padx=10, pady=5)

        self.results_list = tk.Listbox(results_frame, font=("Arial", 9),
                                       height=12, selectmode="single",
                                       activestyle="none")
        scroll = tk.Scrollbar(results_frame, orient="vertical",
                             command=self.results_list.yview)
        self.results_list.configure(yscrollcommand=scroll.set)
        self.results_list.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")
        self.results_list.bind("<Double-Button-1>", lambda e: self.select_book())

        # Qty + Add
        add_frame = tk.Frame(left, bg="white")
        add_frame.pack(fill="x", padx=10, pady=5)
        tk.Label(add_frame, text="Quantity:", font=("Arial", 10), bg="white").pack(side="left")
        self.qty_var = tk.StringVar(value="1")
        tk.Entry(add_frame, textvariable=self.qty_var, width=6,
                font=("Arial", 11)).pack(side="left", padx=5)
        tk.Button(add_frame, text="Add to Quote", command=self.add_to_quote,
                 bg="#1A1A1A", fg="white", font=("Arial", 10, "bold"),
                 relief="flat", padx=15).pack(side="left", padx=5)

        self.status_label = tk.Label(left, text="", font=("Arial", 9),
                                    fg="#C85A00", bg="white")
        self.status_label.pack()

        # RIGHT — Quote panel
        right = tk.Frame(main, bg="white", relief="solid", bd=1, width=400)
        right.pack(side="right", fill="both", padx=(5,0))
        right.pack_propagate(False)

        tk.Label(right, text="CURRENT QUOTE", font=("Arial", 11, "bold"),
                fg="#1A1A1A", bg="white").pack(pady=(10,5))

        # School and ref
        details = tk.Frame(right, bg="white")
        details.pack(fill="x", padx=10)
        tk.Label(details, text="School:", font=("Arial", 9), bg="white").grid(row=0, column=0, sticky="w")
        self.school_var = tk.StringVar()
        tk.Entry(details, textvariable=self.school_var, width=25,
                font=("Arial", 9)).grid(row=0, column=1, padx=5, pady=2)
        tk.Label(details, text="Quote No:", font=("Arial", 9), bg="white").grid(row=1, column=0, sticky="w")
        self.ref_var = tk.StringVar(value="KT-2026-001")
        tk.Entry(details, textvariable=self.ref_var, width=25,
                font=("Arial", 9)).grid(row=1, column=1, padx=5, pady=2)

        # Quote items list
        quote_frame = tk.Frame(right, bg="white")
        quote_frame.pack(fill="both", expand=True, padx=10, pady=5)

        cols = ("Title", "Qty", "Price", "Total")
        self.quote_tree = ttk.Treeview(quote_frame, columns=cols,
                                       show="headings", height=12)
        for col in cols:
            self.quote_tree.heading(col, text=col)
        self.quote_tree.column("Title", width=180)
        self.quote_tree.column("Qty", width=40, anchor="center")
        self.quote_tree.column("Price", width=70, anchor="e")
        self.quote_tree.column("Total", width=80, anchor="e")

        qscroll = tk.Scrollbar(quote_frame, orient="vertical",
                              command=self.quote_tree.yview)
        self.quote_tree.configure(yscrollcommand=qscroll.set)
        self.quote_tree.pack(side="left", fill="both", expand=True)
        qscroll.pack(side="right", fill="y")

        # Remove button
        tk.Button(right, text="Remove Selected", command=self.remove_item,
                 bg="#dc3545", fg="white", font=("Arial", 9),
                 relief="flat", padx=10).pack(pady=2)

        # Grand total
        self.total_label = tk.Label(right, text="GRAND TOTAL: R0.00",
                                   font=("Arial", 12, "bold"),
                                   fg="white", bg="#C85A00")
        self.total_label.pack(fill="x", padx=10, pady=5, ipady=8)

        # Export button
        tk.Button(right, text="EXPORT QUOTE TO EXCEL",
                 command=self.export,
                 bg="#1A1A1A", fg="white",
                 font=("Arial", 11, "bold"),
                 relief="flat", padx=20, pady=8).pack(pady=5)

        self.search_results = []

    def do_search(self):
        query = self.search_var.get().strip()
        if not query:
            return
        self.search_results = search_books(query)
        self.results_list.delete(0, tk.END)
        if not self.search_results:
            self.results_list.insert(tk.END, "  No results found")
            return
        for book in self.search_results:
            isbn = str(book["isbn"]).replace(".0","")
            title = book["title"][:45]
            grade = book["grade"]
            price = book["price"]
            publisher = book["publisher"][:12]
            self.results_list.insert(tk.END,
                f"  {isbn} | {title} | Gr{grade} | R{price} | {publisher}")

    def select_book(self):
        sel = self.results_list.curselection()
        if sel:
            book = self.search_results[sel[0]]
            self.status_label.config(
                text=f"Selected: {book['title'][:40]}")

    def add_to_quote(self):
        sel = self.results_list.curselection()
        if not sel:
            messagebox.showwarning("Select Book",
                                  "Please select a book from the results first.")
            return
        try:
            qty = int(self.qty_var.get())
            if qty <= 0:
                raise ValueError
        except ValueError:
            messagebox.showwarning("Invalid Qty", "Please enter a valid quantity.")
            return

        book = dict(self.search_results[sel[0]])
        book["qty"] = qty
        self.quote_items.append(book)

        title = book["title"][:35]
        price = float(book["price"])
        line = price * qty
        self.quote_tree.insert("", "end",
            values=(title, qty, f"R{price:,.2f}", f"R{line:,.2f}"))
        self.update_total()
        self.status_label.config(
            text=f"✓ Added: {title[:30]} x{qty} = R{line:,.2f}")

    def remove_item(self):
        sel = self.quote_tree.selection()
        if sel:
            idx = self.quote_tree.index(sel[0])
            self.quote_tree.delete(sel[0])
            self.quote_items.pop(idx)
            self.update_total()

    def update_total(self):
        total = sum(float(i["price"]) * int(i["qty"]) for i in self.quote_items)
        self.total_label.config(text=f"GRAND TOTAL: R{total:,.2f}")

    def export(self):
        if not self.quote_items:
            messagebox.showwarning("Empty Quote", "Add books before exporting.")
            return
        school = self.school_var.get().strip() or "School"
        ref_no = self.ref_var.get().strip() or "KT-2026-001"
        filename, total = export_quote(self.quote_items, school, ref_no)
        messagebox.showinfo("Quote Saved",
            f"Quote saved as:\n{filename}\n\nGrand Total: R{total:,.2f}")


if __name__ == "__main__":
    root = tk.Tk()
    app = QuoteTool(root)
    root.mainloop()