import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection

BLACK = "1A1A1A"
ORANGE = "C85A00"
WHITE = "FFFFFF"
LGREY = "F5F5F5"
LORANGE = "FFF3E0"
MGREY = "CCCCCC"

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def bfont(bold=False, color=BLACK, size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)

def bdr():
    s = Side(style="thin", color=MGREY)
    return Border(left=s, right=s, top=s, bottom=s)

conn = sqlite3.connect("books.db")
df = pd.read_sql_query("SELECT isbn, title, grade, subject, language, price, publisher FROM books", conn)
conn.close()

def clean_isbn(isbn):
    try:
        return int(str(isbn).replace(".0","").strip())
    except:
        return str(isbn).replace(".0","").strip()

df["isbn"] = df["isbn"].apply(clean_isbn)
print(f"Loaded {len(df)} books")

def build(demo=False):
    wb = openpyxl.Workbook()
    
    master = wb.active
    master.title = "MASTER"
    master.sheet_state = "hidden"
    
    headers = ["ISBN","TITLE","GRADE","SUBJECT","LANGUAGE","PRICE","PUBLISHER"]
    for col, h in enumerate(headers, 1):
        c = master.cell(row=1, column=col, value=h)
        c.font = bfont(bold=True, color=WHITE)
        c.fill = fill(BLACK)
    
    books = df.head(50) if demo else df
    for row_idx, row in enumerate(books.itertuples(), 2):
        master.cell(row=row_idx, column=1, value=row.isbn)
        master.cell(row=row_idx, column=2, value=row.title)
        master.cell(row=row_idx, column=3, value=str(row.grade))
        master.cell(row=row_idx, column=4, value=str(row.subject))
        master.cell(row=row_idx, column=5, value=str(row.language))
        master.cell(row=row_idx, column=6, value=float(row.price) if row.price else 0)
        master.cell(row=row_idx, column=7, value=str(row.publisher))
    
    master.protection.sheet = True
    master.protection.password = "kalikeng2026master"
    
    ws = wb.create_sheet("QUOTATION")
    
    for col, w in [("A",20),("B",50),("C",8),("D",14),("E",14),("F",12),("G",12),("H",14)]:
        ws.column_dimensions[col].width = w
    
    ws.merge_cells("A1:H1")
    ws["A1"] = "KALIKENG TRADING AND PROJECTS CC"
    ws["A1"].font = bfont(bold=True, size=16, color=WHITE)
    ws["A1"].fill = fill(BLACK)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells("A2:H2")
    ws["A2"] = "Reg No: 2010/007041/23  |  VAT No: 4340257320  |  Tel: 073 223 9762  |  kalikengtrading@gmail.com"
    ws["A2"].font = bfont(size=9, color=ORANGE)
    ws["A2"].fill = fill(BLACK)
    ws["A2"].alignment = Alignment(horizontal="center")
    
    ws.merge_cells("A3:H3")
    ws["A3"] = "B-BBEE LEVEL 1  |  100% BLACK OWNED  |  135% PROCUREMENT RECOGNITION"
    ws["A3"].font = bfont(bold=True, size=9, color=WHITE)
    ws["A3"].fill = fill(BLACK)
    ws["A3"].alignment = Alignment(horizontal="center")
    
    for col in range(1, 9):
        ws.cell(row=4, column=col).fill = fill(ORANGE)
    ws.row_dimensions[4].height = 4
    ws.row_dimensions[5].height = 6
    
    for row, lc, lv, vc in [(6,"A","QUOTATION NO:","B"),(6,"E","DATE:","F"),
                              (7,"A","SCHOOL / CLIENT:","B"),(7,"E","PREPARED BY:","F"),
                              (8,"A","ADDRESS:","B"),(8,"E","PHONE:","F")]:
        ws.row_dimensions[row].height = 18
        lbl = ws[f"{lc}{row}"]
        lbl.value = lv
        lbl.font = bfont(bold=True, color=ORANGE, size=9)
        lbl.fill = fill(LGREY)
        lbl.border = bdr()
        val = ws[f"{vc}{row}"]
        val.font = bfont(size=9)
        val.fill = fill(WHITE)
        val.border = bdr()
        val.protection = Protection(locked=False)
    
    ws.row_dimensions[9].height = 6
    
    for col, hdr in zip(["A","B","C","D","E","F","G","H"],
                        ["ISBN","TITLE","GR","SUBJECT","LANGUAGE","PUBLISHER","UNIT PRICE","QTY"]):
        c = ws[f"{col}10"]
        c.value = hdr
        c.font = bfont(bold=True, color=WHITE, size=9)
        c.fill = fill(BLACK)
        c.alignment = Alignment(horizontal="center")
        c.border = bdr()
    
    ws.column_dimensions["I"].width = 16
    c = ws["I10"]
    c.value = "LINE TOTAL"
    c.font = bfont(bold=True, color=WHITE, size=9)
    c.fill = fill(BLACK)
    c.alignment = Alignment(horizontal="center")
    c.border = bdr()
    ws.row_dimensions[10].height = 18
    
    MAX_ROWS = 10 if demo else 200
    START = 11
    END = START + MAX_ROWS - 1
    
    for r in range(START, END + 1):
        shade = LGREY if r % 2 == 0 else WHITE
        
        ac = ws[f"A{r}"]
        ac.number_format = "0"
        ac.fill = fill(LORANGE)
        ac.font = bfont(size=9)
        ac.border = bdr()
        ac.alignment = Alignment(horizontal="center")
        ac.protection = Protection(locked=False)
        
        ws[f"B{r}"] = f'=IFERROR(VLOOKUP(A{r},MASTER!$A:$G,2,0),"")'
        ws[f"C{r}"] = f'=IFERROR(VLOOKUP(A{r},MASTER!$A:$G,3,0),"")'
        ws[f"D{r}"] = f'=IFERROR(VLOOKUP(A{r},MASTER!$A:$G,4,0),"")'
        ws[f"E{r}"] = f'=IFERROR(VLOOKUP(A{r},MASTER!$A:$G,5,0),"")'
        ws[f"F{r}"] = f'=IFERROR(VLOOKUP(A{r},MASTER!$A:$G,7,0),"")'
        ws[f"G{r}"] = f'=IFERROR(VLOOKUP(A{r},MASTER!$A:$G,6,0),"")'
        
        hc = ws[f"H{r}"]
        hc.fill = fill(LORANGE)
        hc.font = bfont(size=9)
        hc.border = bdr()
        hc.alignment = Alignment(horizontal="center")
        hc.protection = Protection(locked=False)
        
        ws[f"I{r}"] = f'=IFERROR(IF(H{r}="","",G{r}*H{r}),"")'
        
        for col in ["B","C","D","E","F"]:
            c = ws[f"{col}{r}"]
            c.fill = fill(shade)
            c.font = bfont(size=9)
            c.border = bdr()
        
        for col in ["G","I"]:
            c = ws[f"{col}{r}"]
            c.fill = fill(shade)
            c.font = bfont(size=9)
            c.border = bdr()
            c.number_format = 'R#,##0.00'
    
    tr = END + 1
    ws.merge_cells(f"A{tr}:H{tr}")
    tc = ws[f"A{tr}"]
    tc.value = "GRAND TOTAL"
    tc.font = bfont(bold=True, color=WHITE, size=11)
    tc.fill = fill(BLACK)
    tc.alignment = Alignment(horizontal="right", vertical="center")
    tc.border = bdr()
    ws.row_dimensions[tr].height = 22
    
    gt = ws[f"I{tr}"]
    gt.value = f"=IFERROR(SUM(I{START}:I{END}),0)"
    gt.font = bfont(bold=True, color=WHITE, size=11)
    gt.fill = fill(ORANGE)
    gt.border = bdr()
    gt.number_format = 'R#,##0.00'
    
    if demo:
        wr = tr + 2
        ws.merge_cells(f"A{wr}:I{wr}")
        wm = ws[f"A{wr}"]
        wm.value = "DEMO VERSION — 10 lines only | Contact Sechaba: 073 223 9762 | Full version R1,500"
        wm.font = bfont(bold=True, color=WHITE, size=10)
        wm.fill = fill(ORANGE)
        wm.alignment = Alignment(horizontal="center")
    
    ws.protection.sheet = True
    ws.protection.password = "kalikeng2026"
    ws.protection.enable()
    ws.protection.select_unlocked_cells = False
    
    suffix = "DEMO" if demo else "FULL"
    filename = f"Kalikeng_Quote_Tool_{suffix}.xlsx"
    wb.save(filename)
    print(f"Saved: {filename}")

build(demo=True)
build(demo=False)
print("Done!")