import pandas as pd
import sqlite3
import openpyxl
import re

def create_books_db():
    conn = sqlite3.connect("textbooks.db")
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS books (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            publisher TEXT NOT NULL,
            grade TEXT,
            book_type TEXT,
            price REAL NOT NULL
        )
    """)
    conn.commit()
    conn.close()
    print("textbooks.db created")

def add_isbn_column():
    conn = sqlite3.connect("textbooks.db")
    cursor = conn.cursor()
    try:
        cursor.execute("ALTER TABLE books ADD COLUMN isbn TEXT")
        conn.commit()
    except sqlite3.OperationalError:
        print("Column already exists")
    conn.close()

def import_publisher_pricelist(excel_path, publisher_name, header_row, isbn_col, title_col, price_col, grade_col=None, sheet_name=0):
    """
    sheet_name defaults to 0 (first sheet) to keep MML/OUP calls working
    unchanged - Macmillan's multi-sheet import will pass the real sheet
    name explicitly on each call.
    """
    df = pd.read_excel(excel_path, header=header_row, sheet_name=sheet_name)
    df.columns = df.columns.str.strip()
    print(repr(df.columns.tolist()))
    df = df.dropna(subset=[isbn_col, title_col])
    conn = sqlite3.connect("textbooks.db")
    cursor = conn.cursor()

    imported_count = 0
    skipped_count = 0

    for row_number, (_, row) in enumerate(df.iterrows(), start=header_row + 2):
        grade_value = str(row[grade_col]) if grade_col else ""

        try:
            price_value = parse_price(row[price_col], row_context=f"{publisher_name} row {row_number}")
        except ValueError as e:
            print(f"SKIPPED - {e}")
            skipped_count += 1
            continue

        cursor.execute("""
            INSERT INTO books (title, publisher, grade, book_type, price, isbn)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (row[title_col], publisher_name, grade_value, "", row[price_col], str(row[isbn_col])))
        imported_count += 1

    conn.commit()
    conn.close()
    print(f"Imported {imported_count} books from {publisher_name} ({sheet_name}), skipped {skipped_count} bad rows")


conn = sqlite3.connect("textbooks.db")
cursor = conn.cursor()
cursor.execute("DELETE FROM books")
conn.commit()
conn.close()

def find_pricelist_sheet(excel_path, isbn_keywords=None, price_keywords=None, allow_multiple=False):
    """
    Inspects every sheet in a workbook and returns the name(s) of the sheet
    look like real price lists, by scanning the first 15 rows of each
    sheet for header-like keywords.

    allow_multiple=False (default): raises if more than one sheet matches,
    since for most publishers (MML, OUP) that signals genuine ambiguity
    worth stopping on rather than guessing.

    allow_multiple=True: some publishers (Macmillan SA) genuinely ship
    several real price-list sheet in one workbook - one per grade band.
    Set this explicitly, per publisher, once you've confirmed by eye that
    the multiple matches are all real, not an accident.
    """
    if isbn_keywords is None:
        isbn_keywords = ["isbn", "isbn13", "isbn-13"]
    if price_keywords is None:
        price_keywords = ["price", "unit price", "rrp", "cost"]

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    candidates = []

    for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            isbn_hit = False
            price_hit = False

            for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    cell_text = str(cell).strip().lower()
                    if any(kw in cell_text for kw in isbn_keywords):
                        isbn_hit = True
                    if any(kw in cell_text for kw in price_keywords):
                        price_hit = True

            if isbn_hit and price_hit:
                candidates.append(sheet_name)

    wb.close()

    if not candidates:
        raise ValueError(
            f"No sheet in {excel_path} contains both an ISBN-like and "
            f"price-like header - this publisher's layout needs manual inspection."
        )
    if len(candidates) > 1 and not allow_multiple:
        raise ValueError(
            f"Multiple candidate sheets found: {candidates}. "
            f"Confirm which one is the real price list before importing, "
            f"or pass allow_multiple=True if this publisher genuinelly has "
            f"several real price-list sheets."
        )

    return candidates if allow_multiple else candidates[0]

def parse_price(raw_value, row_context=""):
    """
    Converts a real, messy price cell into a float, handling the actual
    formats publishers ship: currency symbols, thousands separators,
    comma-as-decimal, and stray whitespace.

    Raises explicitly rather than returning 0.0 on failure - a wrong
    silent price is worse than a loud crash during import.
    """
    if raw_value is None:
        raise ValueError(f"Empty price cell{f' at {row_context}' if row_context else ''}")
    if isinstance(raw_value, (int, float)):
        return float(raw_value)
        
    text = str(raw_value).strip()
    text = re.sub(r"[A-Za-z\s]", "", text)

    if not text:
        raise ValueError(f"Price cell had no numeric content{f' at {row_context}' if row_context else ''}: {raw_value!r}")
    if "," in text and "." in text:
        text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        return round(float(text), 2)
    except ValueError:
        raise ValueError(
            f"Could not parse price{f' at {row_context}' if row_context else ''}: {raw_value!r} -> cleaned to {text!r}"
        )

    parse_price(row[price_col],row_context=f"row {row_number}")

def import_macmillan_all_sheets(excel_path):
    """
    Macmillan SA's 'MASTER - PRINTED BOOKS' sheet is a real, combined
    list of every printed book across all grades bands - already
    grade-tagged via its own GRADE column. Importing the per-grade
    sheets (GR 1-3, GR 4-6, etc.) separately would duplicate every
    book, since they're subsets of this same master list.
    'MASTER - eBOOKS' is deliberately excluded - schools don't use it.
    """
    import_publisher_pricelist(
        excel_path,
        "Macmillan SA",
        header_row=2,
        isbn_col="ISBN",
        title_col="TITLE",
        price_col="RETAIL PRICE (incl VAT)",
        grade_col="GRADE",
        sheet_name="MASTER - PRINTED BOOKS",
    )

def create_quotes_tables():
    """
    Creates the two real tables a persistent quotation needs:
    - quotes: one row per quotation (the header info)
    - quote_items: one row per book line, linked back to its quote via quote_id

    Run once. Safe to re-run - CREATE TABLE IF NOT EXISTS won't touch
    existing data.
    """
    conn = sqlite3.connect("textbooks.db")
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS quotes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            quote_number TEXT NOT NULL UNIQUE,
            school TEXT NOT NULL,
            prepared_by TEXT NOT NULL,
            date_created TEXT NOT NULL,
            date_updated TEXT NOT NULL
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS quote_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            quote_id INTEGER NOT NULL,
            isbn TEXT NOT NULL,
            title TEXT NOT NULL,
            grade TEXT,
            price REAL NOT NULL,
            qty INTEGER NOT NULL,
            FOREIGN KEY (quote_id) REFERENCES quotes(id) ON DELETE CASCADE
        )
    """)

    cursor.execute("PRAGMA foreign_keys = ON")

    conn.commit()
    conn.close()
    print("quotes and quote_items tables ready")

if __name__ == "__main__":
    create_books_db()
    add_isbn_column()
    create_quotes_tables()

    import_publisher_pricelist(
        "C:/Users/Sechaba/Documents/business/price list 2026-2027/maskewmillerlearningupdatedpricelists20262027/20262027_Grades_0407_Price_list_MML_Hei.xlsx", "Maskew Miller Longman", 11, "ISBN", "TITLE",
        "RRP Price \n1 July 2026 - \n30 June 2027"
    
    )

    import_publisher_pricelist(
        "C:/Users/Sechaba/Documents/business/price list 2026-2027/oxforduniversitypresspricelistseffective1july20263/OUP_Grade_R12_Price_List_202627.xlsx", "Oxford Successful", 6,
        "ISBN", "TITLE", "PRICE", "GRADE"
    )

    import_macmillan_all_sheets(
        "C:/Users/Sechaba/Documents/business/price list 2026-2027/Macmillan SA Retail Price List 2026-2027.xlsx"
    )


  
