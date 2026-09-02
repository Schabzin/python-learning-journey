import openpyxl

excel_path = "C:/Users/Sechaba/Documents/business/price list 2026-2027/Macmillan SA Retail Price List 2026-2027.xlsx"

wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

print("All sheet names:")
for name in wb.sheetnames:
    print(f" {name!r}")

target_sheet = "GR 1-3 (FP)"
ws = wb[target_sheet]

print(f"\nFirst 5 rows of {target_sheet!r} as openpyxl actual reads them:")
for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
    print(f"Row {row_num}: {row}")

wb.close()